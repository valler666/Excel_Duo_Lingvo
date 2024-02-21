from openpyxl.reader.excel import load_workbook

file_eng = 'eng.xlsx'
file_rus = 'rus.xlsx'

xlEnglish = load_workbook(file_eng)
xlRus = load_workbook(file_rus)

sheet_eng = xlEnglish.get_sheet_by_name('URS requirements')
sheet_rus = xlRus.get_sheet_by_name('URS requirements')
for i in range (21,100):
    cell = 'C'+str(i)
    sheet_eng[cell].value = sheet_eng[cell].value + '\n' + sheet_rus[cell].value
    print(cell)
xlEnglish.save(file_eng)





